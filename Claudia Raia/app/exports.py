import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

STATUS_LABELS = {"green": "Verde", "amber": "Amarelo", "red": "Vermelho"}
STATUS_HEX = {"green": "1E9E5A", "amber": "C98A05", "red": "D43F3F"}
STATUS_RGB = {"green": (30, 158, 90), "amber": (201, 138, 5), "red": (212, 63, 63)}


def _pdf_safe(text):
    """Core PDF fonts only support latin-1; replace anything outside that range
    (em-dashes, curly quotes, emoji, etc.) instead of crashing generation."""
    text = str(text)
    return text.encode("latin-1", "replace").decode("latin-1")

HEADER_FILL = PatternFill(start_color="151B2E", end_color="151B2E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(bottom=Side(style="thin", color="E3E7ED"))


def _autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 3, 12), 50)


INVALID_SHEET_CHARS = ["/", "\\", "*", "?", ":", "[", "]"]


def _safe_sheet_title(title):
    for ch in INVALID_SHEET_CHARS:
        title = title.replace(ch, "-")
    return title[:31]


def _write_items_sheet(wb, title, items):
    ws = wb.create_sheet(title=_safe_sheet_title(title))
    headers = ["Status", "Nome", "Categoria", "Responsável", "Métrica", "Valor", "Meta", "Prazo", "Atualizado em", "Descrição"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for it in items:
        ws.append([
            STATUS_LABELS.get(it.status, it.status),
            it.name,
            it.category or "",
            it.owner or "",
            it.metric_label or "",
            it.metric_value if it.metric_value is not None else "",
            it.metric_target if it.metric_target is not None else "",
            it.due_date.strftime("%d/%m/%Y") if it.due_date else "",
            it.updated_at.strftime("%d/%m/%Y %H:%M") if it.updated_at else "",
            it.description or "",
        ])
        row = ws.max_row
        fill = PatternFill(start_color=STATUS_HEX[it.status], end_color=STATUS_HEX[it.status], fill_type="solid")
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    _autosize(ws)
    return ws


def _write_history_sheet(wb, items):
    ws = wb.create_sheet(title="Histórico")
    headers = ["Item", "Tipo", "Status", "Nota", "Atualizado por", "Data"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.freeze_panes = "A2"

    rows = []
    for it in items:
        for h in it.history:
            rows.append((it.name, it.meta["singular"], h))
    rows.sort(key=lambda r: r[2].changed_at, reverse=True)

    for name, kind, h in rows:
        ws.append([
            name, kind, STATUS_LABELS.get(h.status, h.status), h.note or "",
            h.changed_by or "", h.changed_at.strftime("%d/%m/%Y %H:%M"),
        ])
        row = ws.max_row
        fill = PatternFill(start_color=STATUS_HEX[h.status], end_color=STATUS_HEX[h.status], fill_type="solid")
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).font = Font(color="FFFFFF", bold=True)

    _autosize(ws)
    return ws


def build_excel(items_by_type, meta_by_type, all_items):
    """items_by_type: dict entity_type -> list[Item]. If only one type, only that sheet + history is written."""
    wb = Workbook()
    wb.remove(wb.active)

    for entity_type, items in items_by_type.items():
        _write_items_sheet(wb, meta_by_type[entity_type]["plural"], items)

    _write_history_sheet(wb, all_items)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ReportPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 15)
        self.set_text_color(21, 27, 46)
        self.cell(0, 9, "RAG Tracker - Relatorio de Status", ln=1)
        self.set_font("Helvetica", "", 9)
        self.set_text_color(120, 128, 145)
        self.cell(0, 5, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=1)
        self.ln(3)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Página {self.page_no()}", align="C")

    def section_title(self, text):
        self.ln(2)
        self.set_font("Helvetica", "B", 12)
        self.set_text_color(21, 27, 46)
        self.cell(0, 8, _pdf_safe(text), ln=1)
        self.set_draw_color(227, 231, 237)
        self.line(self.get_x(), self.get_y(), self.get_x() + 190, self.get_y())
        self.ln(3)

    def summary_tiles(self, counts):
        self.set_font("Helvetica", "B", 11)
        w = 60
        labels = [("Verde", counts.get("green", 0), (30, 158, 90)),
                  ("Amarelo", counts.get("amber", 0), (201, 138, 5)),
                  ("Vermelho", counts.get("red", 0), (212, 63, 63))]
        x0 = self.get_x()
        y0 = self.get_y()
        for i, (label, val, color) in enumerate(labels):
            self.set_xy(x0 + i * (w + 5), y0)
            self.set_fill_color(*color)
            self.set_text_color(255, 255, 255)
            self.rect(self.get_x(), self.get_y(), w, 20, style="F")
            self.set_xy(x0 + i * (w + 5), y0 + 3)
            self.set_font("Helvetica", "B", 16)
            self.cell(w, 8, str(val), align="C", ln=2)
            self.set_x(x0 + i * (w + 5))
            self.set_font("Helvetica", "", 9)
            self.cell(w, 6, label, align="C")
        self.set_xy(x0, y0 + 25)
        self.set_text_color(0, 0, 0)

    def items_table(self, items):
        self.set_font("Helvetica", "B", 9)
        self.set_fill_color(21, 27, 46)
        self.set_text_color(255, 255, 255)
        col_widths = [22, 55, 35, 35, 33]
        headers = ["Status", "Nome", "Categoria", "Responsável", "Atualizado"]
        for w, h in zip(col_widths, headers):
            self.cell(w, 7, h, border=0, fill=True)
        self.ln()

        self.set_font("Helvetica", "", 8.5)
        for it in items:
            color = STATUS_RGB[it.status]
            self.set_fill_color(*color)
            self.set_text_color(255, 255, 255)
            self.cell(col_widths[0], 6.5, STATUS_LABELS[it.status], fill=True)
            self.set_text_color(30, 30, 30)
            self.set_fill_color(250, 250, 251)
            self.cell(col_widths[1], 6.5, self._trunc(it.name, 34), fill=True)
            self.cell(col_widths[2], 6.5, self._trunc(it.category or "-", 20), fill=True)
            self.cell(col_widths[3], 6.5, self._trunc(it.owner or "-", 20), fill=True)
            self.cell(col_widths[4], 6.5, it.updated_at.strftime("%d/%m/%Y"), fill=True)
            self.ln()
        self.ln(4)

    @staticmethod
    def _trunc(text, n):
        text = _pdf_safe(text)
        return text if len(text) <= n else text[: n - 1] + "."


def build_pdf(title, total_counts, sections):
    """sections: list of (section_title, counts_dict, items_list)"""
    pdf = ReportPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    pdf.section_title(f"Resumo geral - {title}")
    pdf.summary_tiles(total_counts)

    for sec_title, counts, items in sections:
        if not items:
            continue
        pdf.section_title(f"{sec_title} ({len(items)})")
        pdf.items_table(items)

    out = pdf.output()
    return io.BytesIO(bytes(out))
