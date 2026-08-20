"""Shared helpers for bulk-importing Demandas/Projetos/Sistemas from a
spreadsheet (.xlsx or .csv), so users can migrate an existing planilha
instead of retyping everything by hand."""

import csv
import io
from datetime import date, datetime

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_ROWS = 2000

_ACCENTS = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçñÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇÑ", "aaaaaeeeeiiiiooooouuuucnAAAAAEEEEIIIIOOOOOUUUUCN")


def _slug(text):
    text = str(text or "").strip().lower().translate(_ACCENTS)
    return "".join(ch for ch in text if ch.isalnum())


def read_rows(upload):
    """Parses an uploaded .xlsx/.csv into a list of dicts keyed by a
    normalized (accent/case/space-insensitive) version of the header row."""
    filename = (upload.filename or "").lower()
    raw = upload.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError("Arquivo excede o limite de 5 MB.")
    if not raw:
        raise ValueError("Arquivo vazio.")

    if filename.endswith(".xlsx"):
        rows = _read_xlsx(raw)
    elif filename.endswith(".csv"):
        rows = _read_csv(raw)
    else:
        raise ValueError("Formato não suportado. Envie um arquivo .xlsx ou .csv.")

    if len(rows) > MAX_ROWS:
        raise ValueError(f"A planilha tem {len(rows)} linhas; o limite por importação é {MAX_ROWS}.")
    return rows


def _coerce_cell(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if value is None:
        return ""
    return str(value).strip()


def _read_xlsx(raw):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency always installed in prod
        raise ValueError("Suporte a .xlsx indisponível no servidor.") from exc

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Não foi possível ler o arquivo .xlsx: {exc}") from exc

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [_slug(h) for h in header]

    rows = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row = {}
        for key, value in zip(keys, raw_row):
            if key:
                row[key] = _coerce_cell(value)
        rows.append(row)
    return rows


def _read_csv(raw):
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = []
    for raw_row in reader:
        if not raw_row or all((v or "").strip() == "" for v in raw_row.values() if v is not None):
            continue
        row = {}
        for key, value in raw_row.items():
            slug = _slug(key)
            if slug:
                row[slug] = (value or "").strip()
        rows.append(row)
    return rows


def get(row, *aliases, default=""):
    """Fetches the first alias present in a normalized row dict."""
    for alias in aliases:
        slug = _slug(alias)
        value = row.get(slug)
        if value not in (None, ""):
            return value
    return default


def resolve_enum(value, keys, labels=None, default=None):
    """Matches a free-typed spreadsheet value against enum keys or labels,
    ignoring accents/case/spacing (e.g. 'Crítica', 'critica' or 'CRITICA' all
    resolve to the 'critica' key)."""
    if not value:
        return default
    slug = _slug(value)
    for key in keys:
        if _slug(key) == slug:
            return key
    if labels:
        for key, label in labels.items():
            if _slug(label) == slug:
                return key
    return default


def parse_date_flex(value):
    from .utils import parse_date

    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    parsed = parse_date(value)
    if parsed:
        return parsed
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def find_usuario(value):
    from .extensions import db
    from .models import Usuario

    value = (str(value or "")).strip()
    if not value:
        return None
    user = Usuario.query.filter(db.func.lower(Usuario.email) == value.lower()).first()
    if user:
        return user
    return Usuario.query.filter(db.func.lower(Usuario.nome) == value.lower()).first()


def find_by_nome(model, value):
    from .extensions import db

    value = (str(value or "")).strip()
    if not value:
        return None
    return model.query.filter(db.func.lower(model.nome) == value.lower()).first()
