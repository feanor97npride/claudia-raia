"""Shared helpers for spreadsheet output: blank import templates and full
data listings, both as styled .xlsx (mirrors the import feature's format so
a downloaded template can be filled in and re-imported as-is)."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="10162B", end_color="10162B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FILL = PatternFill(start_color="EEF0F4", end_color="EEF0F4", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="6B7280")

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _write_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in ws[letter])
        ws.column_dimensions[letter].width = min(max(length + 3, 12), 42)


def _to_buffer(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template(headers, example_row=None, note=None):
    """A blank spreadsheet with just the header row (styled) plus one
    greyed-out example row, ready to be filled in and re-uploaded via the
    import feature."""
    wb = Workbook()
    ws = wb.active
    if note:
        ws.append([note])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        ws["A1"].font = Font(italic=True, color="6B7280", size=10)
        ws.append([])

    header_row = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = f"A{header_row + 1}"

    if example_row:
        ws.append(example_row)
        for cell in ws[header_row + 1]:
            cell.fill = EXAMPLE_FILL
            cell.font = EXAMPLE_FONT

    _autosize(ws, len(headers))
    return _to_buffer(wb)


def build_listing(headers, rows):
    """A styled spreadsheet with the current data — one row per record."""
    wb = Workbook()
    ws = wb.active
    _write_header(ws, headers)
    for row in rows:
        ws.append(row)
    _autosize(ws, len(headers))
    return _to_buffer(wb)
